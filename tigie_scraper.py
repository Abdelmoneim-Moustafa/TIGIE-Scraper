# ═══════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════
import sys
import io
import time
import argparse
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
#  SCRAPING CORE  (shared by UI and CLI)
# ═══════════════════════════════════════════════════════════════════

_HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

_SESSION = requests.Session()
_SESSION.headers.update(_HTTP_HEADERS)


def _fetch(url: str, retries: int = 3):
    """GET a URL and return a BeautifulSoup, or None on failure."""
    for attempt in range(retries):
        try:
            r = _SESSION.get(url, timeout=20)
            r.raise_for_status()
            return BeautifulSoup(r.text, "lxml")
        except Exception as exc:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                return None


def _parse_historic(soup) -> dict:
    """
    Find the Historic Record table and return the MOST RECENTLY
    published row as a flat dict.

    Table columns expected:
        HS Code | Description | Valid since | Published | Ad-valorem | Unit | M3 Unit Id
    """
    empty = dict(
        hist_hs_code="", hist_description="", hist_valid_since="",
        hist_published="", hist_ad_valorem="", hist_unit="", hist_m3_unit_id=""
    )
    if not soup:
        return empty

    for table in soup.find_all("table"):
        hrow = table.find("tr")
        if not hrow:
            continue
        hdrs = [th.get_text(strip=True).lower() for th in hrow.find_all(["th", "td"])]
        if "ad-valorem" not in hdrs or "published" not in hdrs:
            continue

        # Found the right table — pick the row with the latest Published date
        data_rows = table.find_all("tr")[1:]
        best_cells, best_date = None, None

        for row in data_rows:
            cells = row.find_all(["td", "th"])
            if len(cells) < 4:
                continue
            try:
                pi  = next(i for i, h in enumerate(hdrs) if "published" in h)
                pv  = cells[pi].get_text(strip=True) if pi < len(cells) else ""
                d   = None
                for fmt in ["%b/%d/%Y", "%d/%m/%Y", "%Y-%m-%d"]:
                    try:
                        d = datetime.strptime(pv, fmt)
                        break
                    except ValueError:
                        pass
                if d and (best_date is None or d > best_date):
                    best_date, best_cells = d, cells
            except StopIteration:
                pass

        if best_cells is None and data_rows:
            best_cells = data_rows[-1].find_all(["td", "th"])

        if not best_cells:
            continue

        def gc(col_name):
            try:
                i = next(j for j, h in enumerate(hdrs) if col_name in h)
                return best_cells[i].get_text(strip=True) if i < len(best_cells) else ""
            except StopIteration:
                return ""

        return dict(
            hist_hs_code     = gc("hs code"),
            hist_description = gc("description"),
            hist_valid_since = gc("valid since"),
            hist_published   = gc("published"),
            hist_ad_valorem  = gc("ad-valorem"),
            hist_unit        = gc("unit"),
            hist_m3_unit_id  = gc("m3"),
        )

    return empty


def _extract_aladi_table(soup) -> list:
    """
    Pull every row from the ALADI table in a soup object.

    Table columns expected:
        Country | Ad-valorem | Only | Published | Valid since | Valid until | Code | Quotas
    """
    rows = []
    if not soup:
        return rows

    for table in soup.find_all("table"):
        hrow = table.find("tr")
        if not hrow:
            continue
        hdrs = [th.get_text(strip=True).lower() for th in hrow.find_all(["th", "td"])]
        if "country" not in hdrs or ("code" not in hdrs and "only" not in hdrs):
            continue

        for row in table.find_all("tr")[1:]:
            cells = row.find_all(["td", "th"])
            if len(cells) < 3:
                continue

            def gc(col):
                try:
                    i = next(j for j, h in enumerate(hdrs) if col in h)
                    return cells[i].get_text(strip=True) if i < len(cells) else ""
                except StopIteration:
                    return ""

            entry = dict(
                aladi_country     = gc("country"),
                aladi_ad_valorem  = gc("ad-valorem"),
                aladi_only        = gc("only"),
                aladi_published   = gc("published"),
                aladi_valid_since = gc("valid since"),
                aladi_valid_until = gc("valid until"),
                aladi_code        = gc("code"),
                aladi_quotas      = gc("quotas"),
            )
            if any(entry.values()):
                rows.append(entry)
    return rows


def _fetch_aladi(code: str, date_param: str, retries: int) -> list:
    """Try several URL patterns to get all ALADI rows."""
    for url in [
        f"https://tigie.com.mx/?hs={code}&date={date_param}&tab=aladi",
        f"https://tigie.com.mx/aladi?hs={code}&date={date_param}",
        f"https://tigie.com.mx/?hs={code}&date={date_param}",
    ]:
        soup = _fetch(url, retries)
        rows = _extract_aladi_table(soup)
        if rows:
            return rows
    return []


def scrape_code(code: str, date_param: str, retries: int = 3) -> dict:
    """
    Full scrape for one HS code.

    Returns:
        {
          "code":     str,
          "historic": dict,   # one record (most recent published)
          "aladi":    list,   # all ALADI rows
          "status":   str,    # "OK" or error message
        }
    """
    url  = f"https://tigie.com.mx/?hs={code}&date={date_param}"
    soup = _fetch(url, retries)

    historic = _parse_historic(soup)
    aladi    = _fetch_aladi(code, date_param, retries)

    # Fallback: try to extract ALADI from the main page soup if nothing found
    if not aladi and soup:
        aladi = _extract_aladi_table(soup)

    return {
        "code":     code,
        "historic": historic,
        "aladi":    aladi,
        "status":   "OK",
    }


# ═══════════════════════════════════════════════════════════════════
#  EXCEL BUILDER  (shared by UI and CLI)
# ═══════════════════════════════════════════════════════════════════

def build_excel(results: list) -> bytes:
    """
    Build a styled 3-sheet Excel workbook from scrape results.

    Sheet 1 — Historic Record  : one row per HS code
    Sheet 2 — ALADI            : all ALADI rows (code repeated in col A)
    Sheet 3 — Summary          : counts / run metadata
    """
    wb = Workbook()

    # ── Palette ──────────────────────────────────────────────────
    C_DARK   = "1a1a2e"
    C_NAVY   = "0f3460"
    C_YELLOW = "e2b714"
    C_ALT    = "f1f5f9"

    FILL_DARK   = PatternFill("solid", start_color=C_DARK)
    FILL_NAVY   = PatternFill("solid", start_color=C_NAVY)
    FILL_YELLOW = PatternFill("solid", start_color=C_YELLOW)
    FILL_ALT    = PatternFill("solid", start_color=C_ALT)

    F_HDR  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    F_BODY = Font(name="Arial", size=10)
    F_BOLD = Font(name="Arial", bold=True, size=10)
    F_KEY  = Font(name="Arial", bold=True, size=10, color="000000")
    F_OK   = Font(name="Arial", size=10, color="16a34a")
    F_ERR  = Font(name="Arial", size=10, color="dc2626")

    _side   = Side(style="thin", color="cbd5e1")
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)

    def sh(cell, fill=FILL_DARK):
        cell.font      = F_HDR
        cell.fill      = fill
        cell.border    = _border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def sb(cell, fill=None, key=False, status=None):
        cell.font      = (F_KEY if key else F_BODY)
        cell.border    = _border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:   cell.fill = fill
        if key:    cell.fill = FILL_YELLOW
        if status == "OK":    cell.font = F_OK
        if status == "ERROR": cell.font = F_ERR

    # ── Sheet 1 : Historic Record ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Historic Record"

    H1 = ["HS Code", "Hist HS Code", "Description",
          "Valid Since", "Published", "Ad-Valorem", "Unit", "M3 Unit Id",
          "ALADI Rows", "Status"]
    ws1.append(H1)
    ws1.row_dimensions[1].height = 30
    for ci in range(1, len(H1) + 1):
        sh(ws1.cell(1, ci))

    for ri, r in enumerate(results, 2):
        h    = r["historic"]
        st   = r.get("status", "OK")
        row  = [
            r["code"],
            h.get("hist_hs_code", ""),
            h.get("hist_description", ""),
            h.get("hist_valid_since", ""),
            h.get("hist_published", ""),
            h.get("hist_ad_valorem", ""),
            h.get("hist_unit", ""),
            h.get("hist_m3_unit_id", ""),
            len(r["aladi"]),
            st,
        ]
        ws1.append(row)
        fill = FILL_ALT if ri % 2 == 0 else None
        for ci, _ in enumerate(row, 1):
            c = ws1.cell(ri, ci)
            sb(c, fill=fill, key=(ci == 1),
               status=("OK" if (ci == 10 and st == "OK")
                        else ("ERROR" if ci == 10 else None)))

    for ci, w in enumerate([15, 15, 62, 15, 15, 14, 10, 13, 12, 12], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(H1))}1"

    # ── Sheet 2 : ALADI ───────────────────────────────────────────
    ws2 = wb.create_sheet("ALADI")

    H2 = ["HS Code", "Country", "Ad-Valorem", "Only",
          "Published", "Valid Since", "Valid Until", "ALADI Code", "Quotas"]
    ws2.append(H2)
    ws2.row_dimensions[1].height = 30
    for ci in range(1, len(H2) + 1):
        sh(ws2.cell(1, ci), fill=FILL_NAVY)

    rn = 2
    for r in results:
        for a in r["aladi"]:
            row = [
                r["code"],
                a.get("aladi_country", ""),
                a.get("aladi_ad_valorem", ""),
                a.get("aladi_only", ""),
                a.get("aladi_published", ""),
                a.get("aladi_valid_since", ""),
                a.get("aladi_valid_until", ""),
                a.get("aladi_code", ""),
                a.get("aladi_quotas", ""),
            ]
            ws2.append(row)
            fill = FILL_ALT if rn % 2 == 0 else None
            for ci, _ in enumerate(row, 1):
                sb(ws2.cell(rn, ci), fill=fill, key=(ci == 1))
            rn += 1

    for ci, w in enumerate([15, 20, 16, 62, 15, 15, 15, 13, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(H2))}1"

    # ── Sheet 3 : Summary ─────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "TIGIE Scraper — Run Summary"
    ws3["A1"].font = Font(name="Arial", bold=True, size=14, color=C_DARK)

    stats = [
        ("Total codes processed",      len(results)),
        ("Codes with Historic data",   sum(1 for r in results if r["historic"].get("hist_description"))),
        ("Codes with ALADI data",      sum(1 for r in results if r["aladi"])),
        ("Total ALADI rows",           sum(len(r["aladi"]) for r in results)),
        ("Errors",                     sum(1 for r in results if r.get("status", "OK") != "OK")),
        ("Run date",                   datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(stats, 3):
        ws3.cell(i, 1).value = k
        ws3.cell(i, 1).font  = Font(name="Arial", bold=True, size=11)
        ws3.cell(i, 2).value = v
        ws3.cell(i, 2).font  = Font(name="Arial", size=11)

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════
#  CLI MODE  (python tigie_scraper.py --cli ...)
# ═══════════════════════════════════════════════════════════════════

def _run_cli():
    p = argparse.ArgumentParser(
        description="TIGIE Tariff Scraper — CLI mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python tigie_scraper.py --cli codes.xlsx
  python tigie_scraper.py --cli codes.xlsx --limit 100 --delay 2.0
  python tigie_scraper.py --cli codes.xlsx --start 201 --limit 100
        """,
    )
    p.add_argument("--cli",     required=True, metavar="EXCEL_FILE",
                   help="Input Excel file with a 'Code' column")
    p.add_argument("--date",    default="1/25/2026",
                   help="Date parameter for the URL (default: 1/25/2026)")
    p.add_argument("--delay",   type=float, default=1.5,
                   help="Seconds to wait between requests (default: 1.5)")
    p.add_argument("--retries", type=int,   default=3,
                   help="Max retries per code (default: 3)")
    p.add_argument("--start",   type=int,   default=1,
                   help="Start from this row number (1-indexed, default: 1)")
    p.add_argument("--limit",   type=int,   default=None,
                   help="Max codes to process (default: all)")
    args = p.parse_args()

    df = pd.read_excel(args.cli, dtype={"Code": str})
    df.columns = [c.strip() for c in df.columns]
    if "Code" not in df.columns:
        sys.exit("❌  No 'Code' column found in the file.")

    codes = df["Code"].astype(str).str.strip().tolist()
    codes = codes[args.start - 1:]
    if args.limit:
        codes = codes[:args.limit]

    print(f"\n🚀  Scraping {len(codes)} codes | date={args.date} | delay={args.delay}s\n")

    results = []
    for i, code in enumerate(codes, 1):
        print(f"  [{i:>5}/{len(codes)}]  {code}", end="   ", flush=True)
        try:
            data = scrape_code(code, args.date, args.retries)
            h = data["historic"].get("hist_description")
            print(f"hist={'✓' if h else '✗'}  aladi={len(data['aladi'])}")
        except Exception as exc:
            data = {
                "code": code,
                "historic": {k: "" for k in [
                    "hist_hs_code", "hist_description", "hist_valid_since",
                    "hist_published", "hist_ad_valorem", "hist_unit", "hist_m3_unit_id",
                ]},
                "aladi":  [],
                "status": f"ERROR: {exc}",
            }
            print(f"ERROR: {exc}")
        results.append(data)
        time.sleep(args.delay)

    out = f"tigie_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with open(out, "wb") as f:
        f.write(build_excel(results))

    n_hist  = sum(1 for r in results if r["historic"].get("hist_description"))
    n_aladi = sum(1 for r in results if r["aladi"])
    n_err   = sum(1 for r in results if r.get("status", "OK") != "OK")
    print(f"\n✅  Done — {len(results)} codes | {n_hist} historic | {n_aladi} with ALADI | {n_err} errors")
    print(f"📄  Saved → {out}\n")


# ═══════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════

def _run_streamlit():
    import streamlit as st  # imported here so CLI works without streamlit

    st.set_page_config(
        page_title="TIGIE Scraper",
        page_icon="📦",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── CSS ──────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .header-box{
        background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);
        padding:2rem 2.5rem;border-radius:14px;margin-bottom:1.5rem;text-align:center;
    }
    .header-box h1{color:#e2b714;margin:0;font-size:2.2rem;letter-spacing:-0.5px;}
    .header-box p {color:#a0aec0;margin:.5rem 0 0;font-size:1rem;}
    .kpi{
        background:#1e293b;border:1px solid #334155;border-radius:10px;
        padding:1.1rem;text-align:center;
    }
    .kpi .v{font-size:2rem;font-weight:700;color:#e2b714;}
    .kpi .l{font-size:.8rem;color:#94a3b8;margin-top:.2rem;}
    .tip{
        background:#0c1a27;border-left:4px solid #e2b714;
        padding:.8rem 1rem;border-radius:0 8px 8px 0;
        font-size:.9rem;color:#cbd5e1;margin:.6rem 0;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Header ───────────────────────────────────────────────────
    st.markdown("""
    <div class="header-box">
      <h1>📦 TIGIE Tariff Scraper</h1>
      <p>Automated extraction of Historic Records &amp; ALADI data from tigie.com.mx</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar ──────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### ⚙️ Settings")
        delay       = st.slider("Delay between requests (sec)", 0.5, 5.0, 1.5, 0.5)
        max_retries = st.number_input("Max retries per code", 1, 5, 3)
        date_param  = st.text_input("Date parameter", value="1/25/2026")

        st.markdown("---")
        st.markdown("### 📖 How to use")
        st.markdown("""
1. Upload your Excel file  
2. (Optional) filter by priority  
3. Choose row range to process  
4. Click **Start Scraping**  
5. Download the Excel result  

**Data per code:**  
- Historic Record → most recent Published, Ad-valorem, Description  
- ALADI → all country/agreement rows  
""")
        st.markdown("---")
        st.markdown("### 🔗 URL pattern")
        st.code("https://tigie.com.mx/?hs={CODE}&date={DATE}", language="")

    # ── File upload ──────────────────────────────────────────────
    st.markdown("### 📁 Upload your codes file")
    uploaded = st.file_uploader(
        "Excel with a `Code` column (optionally `Link`, `Priority`)",
        type=["xlsx", "xls"],
    )

    if not uploaded:
        st.markdown("""
        <div class="tip">
        📌 Upload your <code>Copy_of_codes__Links.xlsx</code> to get started.
        The file must have a <code>Code</code> column with HS codes like <code>85011010</code>.
        </div>
        """, unsafe_allow_html=True)
        st.markdown("### 📋 What gets extracted")
        st.dataframe(pd.DataFrame({
            "Sheet":   ["Historic Record","Historic Record","ALADI","ALADI"],
            "Column":  ["Published","Ad-Valorem","Country","ALADI Code"],
            "Example": ["Jun/07/2022","Ex.","Argentina","ACE6"],
            "Notes":   [
                "Most recent published date",
                "Tariff rate / type",
                "Trading partner country",
                "Agreement identifier",
            ],
        }), use_container_width=True, hide_index=True)
        return

    # ── Parse upload ─────────────────────────────────────────────
    df_in = pd.read_excel(uploaded, dtype={"Code": str})
    df_in.columns = [c.strip() for c in df_in.columns]

    if "Code" not in df_in.columns:
        st.error("❌ No `Code` column found. Please check your file.")
        return

    df_in["Code"] = df_in["Code"].astype(str).str.strip()
    total = len(df_in)

    # Priority filter
    if "Priority" in df_in.columns:
        priorities = sorted(df_in["Priority"].dropna().unique().tolist())
        sel = st.multiselect("Filter by Priority (leave empty = all)",
                             priorities, default=priorities)
        if sel:
            df_in = df_in[df_in["Priority"].isin(sel)]

    # Row-range selector
    c1, c2 = st.columns(2)
    with c1:
        start_i = st.number_input("Start from row #", 1, len(df_in), 1)
    with c2:
        end_i   = st.number_input("End at row #", 1, len(df_in),
                                  min(100, len(df_in)))

    df_batch = df_in.iloc[start_i - 1 : end_i].reset_index(drop=True)

    # KPI row
    k1, k2, k3 = st.columns(3)
    for col, val, lbl in [
        (k1, total,           "Total codes in file"),
        (k2, len(df_in),      "After priority filter"),
        (k3, len(df_batch),   "In current batch"),
    ]:
        col.markdown(
            f'<div class="kpi"><div class="v">{val:,}</div>'
            f'<div class="l">{lbl}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown(
        '<div class="tip">⚡ Tip: Process in batches of 100–200 codes '
        'to avoid rate-limiting.</div>',
        unsafe_allow_html=True,
    )

    with st.expander("Preview batch"):
        st.dataframe(df_batch.head(20), use_container_width=True)

    # ── Scrape button ────────────────────────────────────────────
    if not st.button("🚀 Start Scraping", type="primary", use_container_width=True):
        return

    codes = df_batch["Code"].tolist()
    n     = len(codes)

    prog  = st.progress(0)
    stxt  = st.empty()
    ltbl  = st.empty()

    m1, m2, m3, m4 = st.columns(4)
    mm = {k: c.empty() for k, c in zip(
        ["done", "hist", "aladi", "err"],
        [m1, m2, m3, m4]
    )}

    results   = []
    done      = n_hist = n_aladi = n_err = 0

    for code in codes:
        stxt.markdown(f"⏳ Scraping **{code}** ({done+1}/{n})…")
        try:
            data = scrape_code(code, date_param, max_retries)
            if data["historic"].get("hist_description"): n_hist  += 1
            if data["aladi"]:                             n_aladi += 1
        except Exception as exc:
            data = {
                "code": code,
                "historic": {k: "" for k in [
                    "hist_hs_code","hist_description","hist_valid_since",
                    "hist_published","hist_ad_valorem","hist_unit","hist_m3_unit_id",
                ]},
                "aladi":  [],
                "status": f"ERROR: {str(exc)[:80]}",
            }
            n_err += 1

        results.append(data)
        done += 1
        prog.progress(done / n)

        mm["done"].metric("✅ Done",        done)
        mm["hist"].metric("📜 Historic",    n_hist)
        mm["aladi"].metric("🌎 ALADI",      n_aladi)
        mm["err"].metric("❌ Errors",       n_err)

        # Live preview (last 5 rows)
        ltbl.dataframe(
            pd.DataFrame([{
                "Code":        r["code"],
                "Description": r["historic"].get("hist_description","")[:55],
                "Published":   r["historic"].get("hist_published",""),
                "Ad-valorem":  r["historic"].get("hist_ad_valorem",""),
                "ALADI rows":  len(r["aladi"]),
                "Status":      r.get("status","OK"),
            } for r in results[-5:]]),
            use_container_width=True,
        )
        time.sleep(delay)

    stxt.markdown("✅ **Scraping complete!**")

    # ── Download ─────────────────────────────────────────────────
    excel_bytes = build_excel(results)
    fname = f"tigie_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    st.success(
        f"🎉 {done} codes | {n_hist} with historic | "
        f"{n_aladi} with ALADI | {n_err} errors"
    )
    st.download_button(
        label="⬇️ Download Excel",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    # ── Charts ───────────────────────────────────────────────────
    st.markdown("### 📊 Results Analysis")
    df_res = pd.DataFrame([{
        "Code":         r["code"],
        "Has Historic": bool(r["historic"].get("hist_description")),
        "ALADI Rows":   len(r["aladi"]),
        "Ad-valorem":   r["historic"].get("hist_ad_valorem",""),
        "Status":       r.get("status","OK"),
    } for r in results])

    ca, cb = st.columns(2)
    with ca:
        st.markdown("**Data coverage**")
        st.bar_chart(pd.DataFrame({
            "Count": {
                "Has Historic": sum(df_res["Has Historic"]),
                "Has ALADI":    sum(df_res["ALADI Rows"] > 0),
                "Has Both":     sum(df_res["Has Historic"] & (df_res["ALADI Rows"] > 0)),
                "No Data":      sum(~df_res["Has Historic"] & (df_res["ALADI Rows"] == 0)),
            }
        }))
    with cb:
        st.markdown("**Top 10 codes by ALADI rows**")
        st.dataframe(
            df_res.nlargest(10, "ALADI Rows")[["Code","ALADI Rows"]],
            use_container_width=True,
        )

    st.markdown("**Ad-valorem distribution**")
    st.bar_chart(df_res["Ad-valorem"].value_counts().head(15).rename("Count"))


# ═══════════════════════════════════════════════════════════════════
#  ENTRY POINT  — decide CLI vs Streamlit
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if "--cli" in sys.argv:
        _run_cli()
    else:
        # When launched via `streamlit run`, __name__ is not __main__,
        # so Streamlit picks up the module-level call below.
        pass

# Module-level Streamlit call (executed when `streamlit run tigie_scraper.py`)
try:
    import streamlit as _st_check
    _in_streamlit = hasattr(_st_check, "_is_running_with_streamlit")
except ImportError:
    _in_streamlit = False

# Always try to run the Streamlit UI when not in CLI mode
if "--cli" not in sys.argv:
    try:
        _run_streamlit()
    except Exception:
        pass  # Silently skip if not in a Streamlit context
